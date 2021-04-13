from rest_framework import status
from rest_framework import views
from rest_framework.authentication import TokenAuthentication
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from core.models import Account, Integration, Transactions

from account import serializers

from rest_framework.renderers import JSONRenderer

class AccountTransactionsViewSet(views.APIView):
    """To list quickbooks transactions"""
    authentication_classes = (TokenAuthentication,)
    permission_classes = (IsAuthenticated,)
    
    def get_exported_data(self, queryset):
        """This function is used to sent the transactions data in xlsx format file"""
        from datetime import datetime
        from datetime import timedelta
        from openpyxl import Workbook
        from django.http import HttpResponse

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        response['Content-Disposition'] = 'attachment; filename={date}-transactions.xlsx'.format(
            date=datetime.now().strftime('%Y-%m-%d'),
        )
        workbook = Workbook()
        
        # Get active worksheet/tab
        worksheet = workbook.active
        worksheet.title = 'Transactions'

        # Define the titles for columns
        columns = [
            'ID',
            'Account Name',
            'Account ID',
            'Amount',
            'Type',
            'Date',
            'Integration Name'
        ]
        row_num = 1

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

        # Iterate through all transactions
        for transaction in queryset:
            row_num += 1
            
            # Define the data for each cell in the row 
            row = [
                transaction.id,
                transaction.account_name,
                transaction.account_id,
                abs(transaction.amount),
                'Credit' if transaction.type == 'CR' else 'Debit',
                datetime.strptime(str(transaction.date), "%Y-%m-%d").strftime("%d, %b %Y"),
                transaction.integration_name.capitalize()
            ]
            
            # Assign the data for each cell of the row 
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        workbook.save(response)

        return response


    def get(self, request, *args, **kwargs):
        """Custom GET endpoint to return transactions data"""
        query_filters = []
        query_filters_data = []

        # Filtering objects
        # filter for current logged in user
        query_filters.append(f"ca.user_id = %s")
        query_filters_data.append(self.request.user.id)

        if 'integration_id' in request.query_params:
            query_filters.append(f"ci.id = %s")
            query_filters_data.append(request.query_params['integration_id'])
        if 'fromDate' in request.query_params:
            query_filters.append(f"ct.date >= %s")
            query_filters_data.append(request.query_params['fromDate'])
        if 'toDate' in request.query_params:
            query_filters.append(f"ct.date <= %s")
            query_filters_data.append(request.query_params['toDate'])
        if 'date' in request.query_params:
            query_filters.append(f"ct.date = %s")
            query_filters_data.append(request.query_params['date'])
        if 'type' in request.query_params:
            query_filters.append(f"ct.type = %s")
            query_filters_data.append(request.query_params['type'])
        
        
        # Applying orders
        order = {
            'account_name': 'ct.account_name',
            'account_id': 'ct.account_id',
            'amount': 'ct.amount',
            'date': 'ct.date',
            'type': 'ct.type',
            'integration_id': 'ci.id',
            'integration_name': 'ci.name',
        }
        order_by = order[request.query_params['order_by']] if 'order_by' in request.query_params and order.get(request.query_params['order_by']) is not None else 'ct.date'
        order_direction = request.query_params['order_direction'].upper() if 'order_direction' in request.query_params and request.query_params['order_direction'].upper() in ['ASC', 'DESC'] else 'DESC'

        # Applying pagination
        page = 1 if 'page' not in request.query_params or int(request.query_params['page']) <= 0 else int(request.query_params['page'])
        per_page = 15 if 'per_page' not in request.query_params or int(request.query_params['per_page']) <= 0 else int(request.query_params['per_page'])
        
        # Preparing SQL to get total records
        queryset = Account.objects.raw(f"""SELECT COUNT(*) AS id
                                                FROM core_transactions ct
                                                INNER JOIN core_account ca ON ct.core_account_id = ca.id
                                                INNER JOIN core_integration ci ON ca.integration_id = ci.id
                                                WHERE {' AND '.join(query_filters)}""", query_filters_data)
        total_records = queryset[0].__dict__['id']

        # Preparing SQL to get records
        sql = f"""SELECT ct.id, ct.account_name, ct.account_id, ct.amount, ct.date, ct."type", ci.id AS integration_id, ci.name AS integration_name,
                COUNT(*) over() AS total_count
                FROM core_transactions ct
                INNER JOIN core_account ca ON ct.core_account_id = ca.id
                INNER JOIN core_integration ci ON ca.integration_id = ci.id
                WHERE {' AND '.join(query_filters)}
                ORDER BY {order_by} {order_direction}"""
        limit = f""" LIMIT {int(per_page)} OFFSET {(int(page)-1)*int(per_page)}"""
        if 'download' in request.query_params and request.query_params['download'] == '1':
            queryset = Account.objects.raw(sql, query_filters_data)
            return self.get_exported_data(queryset)
        
        # Adding limit for paginated response
        sql += limit
        queryset = Account.objects.raw(sql, query_filters_data)

        serializer = serializers.TransactionsSerializer(list(queryset), many=True)
        data = {}
        data['total'] = total_records
        data['page'] = page
        data['per_page'] = per_page
        import math
        data['total_page'] = math.ceil(total_records/per_page)
        data['data'] = serializer.data

        return Response(data)

class AccountChartsViewSet(views.APIView):
    """To return data for charts"""
    authentication_classes = (TokenAuthentication,)
    permission_classes = (IsAuthenticated,)
    
    def get_total_deposits(self, user_id):
        """To return total deposits"""
        query_filters = []
        query_filters_data = []

        # Filtering objects
        # filter for current logged in user
        query_filters.append(f"ca.user_id = %s")
        query_filters_data.append(user_id)
        # filter for deposits
        query_filters.append(f"ct.type = %s")
        query_filters_data.append('CR')

        sql = f"""SELECT 1 AS id, ROUND(SUM(ct.amount)::NUMERIC, 2) AS total_deposits
                FROM core_transactions ct
                INNER JOIN core_account ca ON ca.id = ct.core_account_id
                WHERE {' AND '.join(query_filters)}"""
            
        queryset = Account.objects.raw(sql, query_filters_data)
        serializer = serializers.ChartSerializer(list(queryset), many=True)
        data = {}
        data['data'] = serializer.data

        return data
    
    def get_data_by_account(self, user_id):
        """To return total amount by account name"""
        query_filters = []
        query_filters_data = []

        # Filtering objects
        # filter for current logged in user
        query_filters.append(f"ca.user_id = %s")
        query_filters_data.append(user_id)

        sql = f"""SELECT 1 AS id, ROUND(SUM(ct.amount)::NUMERIC, 2) AS total_amount, ct.account_name
                FROM core_transactions ct
                INNER JOIN core_account ca ON ca.id = ct.core_account_id
                WHERE {' AND '.join(query_filters)}
                GROUP BY ct.account_name
                ORDER BY ct.account_name"""
            
        queryset = Account.objects.raw(sql, query_filters_data)
        serializer = serializers.ChartSerializer(list(queryset), many=True)
        data = {}
        data['data'] = serializer.data

        return data

    def get_data_by_date(self, user_id):
        """To return transactions data date wise"""
        query_filters = []
        query_filters_data = []

        # Filtering objects
        # filter for current logged in user
        query_filters.append(f"ca.user_id = %s")
        query_filters_data.append(user_id)

        sql = f"""SELECT 1 AS id, DATE, sum(total_amount) over (order by date) AS total_amount
                FROM
                    (SELECT ROUND(SUM(ct.amount)::NUMERIC, 2) AS total_amount, ct.date
                    FROM core_transactions ct
                    INNER JOIN core_account ca ON ca.id = ct.core_account_id
                    WHERE {' AND '.join(query_filters)}
                    GROUP BY ct.date
                    ORDER BY ct.date) AS a"""
            
        queryset = Account.objects.raw(sql, query_filters_data)
        serializer = serializers.ChartSerializer(list(queryset), many=True)
        data = {}
        data['data'] = serializer.data

        return data
    
    def get_data_by_integration(self, user_id):
        """To return transactions data date wise"""
        query_filters = []
        query_filters_data = []

        # Filtering objects
        # filter for current logged in user
        query_filters.append(f"ca.user_id = %s")
        query_filters_data.append(user_id)

        sql = f"""SELECT 1 AS id, ci.name AS integration_name, 
                SUM(CASE WHEN ct."type" = 'CR' THEN ct.amount::NUMERIC ELSE 0 END) AS credit,
                SUM(CASE WHEN ct."type" = 'DE' THEN ABS(ct.amount)::NUMERIC ELSE 0 END) AS debit

                FROM core_transactions ct
                INNER JOIN core_account ca ON ca.id = ct.core_account_id
                INNER JOIN core_integration ci ON ca.integration_id = ci.id
                WHERE {' AND '.join(query_filters)}
                GROUP BY ci.name"""
            
        queryset = Account.objects.raw(sql, query_filters_data)
        serializer = serializers.ChartSerializer(list(queryset), many=True)
        data = {}
        data['data'] = serializer.data

        return data

    def get(self, request, *args, **kwargs):
        """Custom GET endpoint to return transactions data"""

        chart_type = 1
        if 'chart_type' in request.query_params:
            chart_type = request.query_params['chart_type']
        switcher = {
                        1: self.get_data_by_date,
                        2: self.get_total_deposits,
                        3: self.get_data_by_account,
                        4: self.get_data_by_integration
                    }
        get_chart_data = switcher.get(int(chart_type))
        return Response(get_chart_data(self.request.user.id))

        